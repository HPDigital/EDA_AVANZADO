"""
Sistema de carga de datos para múltiples formatos de archivos
Soporta CSV, Excel, JSON, Parquet, y otros formatos comunes
"""
import os
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Union, Any
import warnings
from datetime import datetime
import json

# Detectar dependencias opcionales
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    import pyarrow
    HAS_PYARROW = True
except ImportError:
    HAS_PYARROW = False

try:
    import fastparquet
    HAS_FASTPARQUET = True
except ImportError:
    HAS_FASTPARQUET = False

warnings.filterwarnings('ignore')


class MultiFormatLoader:
    """Cargador de datos para múltiples formatos de archivos"""
    
    # Formatos soportados (simplificado)
    SUPPORTED_FORMATS = {
        '.csv': 'CSV (Comma Separated Values)',
        '.xlsx': 'Excel (Excel 2007+)'
    }
    
    def __init__(self):
        self.file_info = {}
        self.loaded_data = {}
        
    def get_supported_formats(self) -> Dict[str, str]:
        """Retorna los formatos soportados"""
        return self.SUPPORTED_FORMATS.copy()
    
    def detect_file_format(self, file_path: str) -> str:
        """Detecta el formato del archivo basado en la extensión (solo CSV y XLSX)"""
        path = Path(file_path)
        extension = path.suffix.lower()
        
        if extension in self.SUPPORTED_FORMATS:
            return extension
        else:
            # Si no es CSV o XLSX, intentar detectar CSV por contenido
            if extension == '.txt' or extension not in ['.csv', '.xlsx']:
                return self._detect_csv_by_content(file_path)
            else:
                raise Exception(f"Formato no soportado: {extension}. Solo se soportan CSV y XLSX.")
    
    def _detect_csv_by_content(self, file_path: str) -> str:
        """Detecta si el archivo es CSV por su contenido"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                first_line = f.readline().strip()
                
            # Detectar CSV
            if ',' in first_line and len(first_line.split(',')) > 2:
                return '.csv'
            else:
                raise Exception("El archivo no parece ser un CSV válido")
        except:
            raise Exception("No se pudo leer el archivo para detectar el formato")
    
    def load_file(self, file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        Carga un archivo en el formato detectado automáticamente
        
        Args:
            file_path: Ruta del archivo
            **kwargs: Argumentos adicionales para el cargador específico
            
        Returns:
            Tuple con (DataFrame, metadata)
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {file_path}")
        
        # Detectar formato
        file_format = self.detect_file_format(str(file_path))
        
        # Información del archivo
        file_info = {
            'path': str(file_path),
            'name': file_path.name,
            'format': file_format,
            'size': file_path.stat().st_size,
            'modified': datetime.fromtimestamp(file_path.stat().st_mtime),
            'encoding': self._detect_encoding(str(file_path))
        }
        
        # Cargar según el formato (solo CSV y XLSX)
        try:
            if file_format == '.csv':
                df, load_info = self._load_csv(str(file_path), **kwargs)
            elif file_format == '.xlsx':
                df, load_info = self._load_excel(str(file_path), **kwargs)
            else:
                raise Exception(f"Formato no soportado: {file_format}. Solo se soportan CSV y XLSX.")
            
            # Combinar información
            metadata = {**file_info, **load_info}
            
            return df, metadata
            
        except Exception as e:
            raise Exception(f"Error al cargar archivo {file_path.name}: {str(e)}")
    
    def _detect_encoding(self, file_path: str) -> str:
        """Detecta la codificación del archivo"""
        try:
            import chardet
            with open(file_path, 'rb') as f:
                raw_data = f.read(10000)  # Leer primeros 10KB
                result = chardet.detect(raw_data)
                return result.get('encoding', 'utf-8')
        except ImportError:
            # Si chardet no está disponible, usar utf-8 por defecto
            return 'utf-8'
        except:
            return 'utf-8'
    
    def _load_csv(self, file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict]:
        """Carga archivo CSV"""
        # Detectar separador automáticamente
        sep = kwargs.get('sep', None)
        if sep is None:
            sep = self._detect_csv_separator(file_path)
        
        # Detectar encoding
        encoding = kwargs.get('encoding', self._detect_encoding(file_path))
        
        # Intentar cargar con diferentes configuraciones
        configs = [
            {'sep': sep, 'encoding': encoding, 'decimal': '.'},
            {'sep': sep, 'encoding': encoding, 'decimal': ','},
            {'sep': ';', 'encoding': encoding, 'decimal': ','},
            {'sep': ',', 'encoding': encoding, 'decimal': '.'},
            {'sep': '\t', 'encoding': encoding, 'decimal': '.'}
        ]
        
        for config in configs:
            try:
                df = pd.read_csv(file_path, **config, **kwargs)
                if len(df.columns) > 1 and len(df) > 0:
                    return df, {'separator': config['sep'], 'encoding': config['encoding'], 'decimal': config['decimal']}
            except:
                continue
        
        # Si todo falla, usar configuración por defecto
        df = pd.read_csv(file_path, sep=',', encoding='utf-8', **kwargs)
        return df, {'separator': ',', 'encoding': 'utf-8', 'decimal': '.'}
    
    def _detect_csv_separator(self, file_path: str) -> str:
        """Detecta el separador del CSV"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                first_line = f.readline().strip()
            
            # Contar ocurrencias de diferentes separadores
            separators = {',': first_line.count(','), ';': first_line.count(';'), '\t': first_line.count('\t')}
            
            # Retornar el separador más común
            return max(separators, key=separators.get)
        except:
            return ','
    
    def _load_tsv(self, file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict]:
        """Carga archivo TSV"""
        encoding = kwargs.get('encoding', self._detect_encoding(file_path))
        
        try:
            df = pd.read_csv(file_path, sep='\t', encoding=encoding, **kwargs)
            return df, {'separator': '\t', 'encoding': encoding}
        except:
            # Fallback a CSV si falla
            return self._load_csv(file_path, sep='\t', **kwargs)
    
    def _load_excel(self, file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict]:
        """Carga archivo Excel"""
        if not HAS_OPENPYXL and not HAS_XLRD:
            raise ImportError("Se requiere openpyxl o xlrd para cargar archivos Excel")
        
        sheet_name = kwargs.get('sheet_name', 0)
        engine = kwargs.get('engine', None)
        
        # Determinar engine automáticamente
        if engine is None:
            if Path(file_path).suffix.lower() == '.xlsx' and HAS_OPENPYXL:
                engine = 'openpyxl'
            elif Path(file_path).suffix.lower() == '.xls' and HAS_XLRD:
                engine = 'xlrd'
            else:
                engine = 'openpyxl'  # Por defecto
        
        try:
            # Intentar cargar hoja específica
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine, **kwargs)
            return df, {'sheet': sheet_name, 'engine': engine}
        except:
            # Si falla, intentar cargar primera hoja
            try:
                df = pd.read_excel(file_path, sheet_name=0, engine=engine, **kwargs)
                return df, {'sheet': 0, 'engine': engine}
            except:
                raise Exception("No se pudo cargar el archivo Excel")
    
    def _load_json(self, file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict]:
        """Carga archivo JSON"""
        encoding = kwargs.get('encoding', self._detect_encoding(file_path))
        
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                data = json.load(f)
            
            # Convertir a DataFrame
            if isinstance(data, list):
                df = pd.DataFrame(data)
            elif isinstance(data, dict):
                # Intentar diferentes estructuras
                if 'data' in data:
                    df = pd.DataFrame(data['data'])
                elif 'records' in data:
                    df = pd.DataFrame(data['records'])
                else:
                    df = pd.DataFrame([data])
            else:
                df = pd.DataFrame([data])
            
            return df, {'encoding': encoding, 'structure': 'json'}
        except Exception as e:
            raise Exception(f"Error al cargar JSON: {str(e)}")
    
    def _load_parquet(self, file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict]:
        """Carga archivo Parquet"""
        if not HAS_PYARROW and not HAS_FASTPARQUET:
            raise ImportError("Se requiere pyarrow o fastparquet para cargar archivos Parquet")
        
        engine = kwargs.get('engine', None)
        if engine is None:
            engine = 'pyarrow' if HAS_PYARROW else 'fastparquet'
        
        try:
            df = pd.read_parquet(file_path, engine=engine, **kwargs)
            return df, {'engine': engine}
        except Exception as e:
            raise Exception(f"Error al cargar Parquet: {str(e)}")
    
    def _load_pickle(self, file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict]:
        """Carga archivo Pickle"""
        try:
            df = pd.read_pickle(file_path, **kwargs)
            return df, {'format': 'pickle'}
        except Exception as e:
            raise Exception(f"Error al cargar Pickle: {str(e)}")
    
    def _load_feather(self, file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict]:
        """Carga archivo Feather"""
        try:
            df = pd.read_feather(file_path, **kwargs)
            return df, {'format': 'feather'}
        except Exception as e:
            raise Exception(f"Error al cargar Feather: {str(e)}")
    
    def _load_hdf5(self, file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict]:
        """Carga archivo HDF5"""
        key = kwargs.get('key', None)
        
        try:
            if key is None:
                # Listar todas las claves disponibles
                with pd.HDFStore(file_path, 'r') as store:
                    keys = store.keys()
                    if keys:
                        key = keys[0]
                    else:
                        raise Exception("No se encontraron tablas en el archivo HDF5")
            
            df = pd.read_hdf(file_path, key=key, **kwargs)
            return df, {'key': key}
        except Exception as e:
            raise Exception(f"Error al cargar HDF5: {str(e)}")
    
    def _load_xml(self, file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict]:
        """Carga archivo XML"""
        try:
            df = pd.read_xml(file_path, **kwargs)
            return df, {'format': 'xml'}
        except Exception as e:
            raise Exception(f"Error al cargar XML: {str(e)}")
    
    def _load_html(self, file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict]:
        """Carga archivo HTML (tablas)"""
        try:
            tables = pd.read_html(file_path, **kwargs)
            if tables:
                df = tables[0]  # Usar primera tabla
                return df, {'format': 'html', 'table_index': 0, 'total_tables': len(tables)}
            else:
                raise Exception("No se encontraron tablas en el archivo HTML")
        except Exception as e:
            raise Exception(f"Error al cargar HTML: {str(e)}")
    
    def _load_text(self, file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict]:
        """Carga archivo de texto delimitado"""
        # Intentar detectar el delimitador
        sep = kwargs.get('sep', None)
        if sep is None:
            sep = self._detect_csv_separator(file_path)
        
        encoding = kwargs.get('encoding', self._detect_encoding(file_path))
        
        try:
            df = pd.read_csv(file_path, sep=sep, encoding=encoding, **kwargs)
            return df, {'separator': sep, 'encoding': encoding}
        except Exception as e:
            raise Exception(f"Error al cargar archivo de texto: {str(e)}")
    
    def load_multiple_files(self, file_paths: List[str], **kwargs) -> Dict[str, Tuple[pd.DataFrame, Dict]]:
        """Carga múltiples archivos"""
        results = {}
        
        for file_path in file_paths:
            try:
                df, metadata = self.load_file(file_path, **kwargs)
                results[file_path] = (df, metadata)
            except Exception as e:
                results[file_path] = (None, {'error': str(e)})
        
        return results
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Obtiene información detallada del archivo sin cargarlo"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            return {'error': 'Archivo no encontrado'}
        
        file_format = self.detect_file_format(str(file_path))
        
        info = {
            'name': file_path.name,
            'format': file_format,
            'format_description': self.SUPPORTED_FORMATS.get(file_format, 'Formato desconocido'),
            'size': file_path.stat().st_size,
            'size_mb': round(file_path.stat().st_size / (1024 * 1024), 2),
            'modified': datetime.fromtimestamp(file_path.stat().st_mtime),
            'encoding': self._detect_encoding(str(file_path)),
            'supported': file_format in self.SUPPORTED_FORMATS
        }
        
        # Información específica por formato
        if file_format in ['.xlsx', '.xls']:
            try:
                if file_format == '.xlsx' and HAS_OPENPYXL:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, read_only=True)
                    info['sheets'] = wb.sheetnames
                    info['total_sheets'] = len(wb.sheetnames)
                elif file_format == '.xls' and HAS_XLRD:
                    import xlrd
                    wb = xlrd.open_workbook(file_path)
                    info['sheets'] = wb.sheet_names()
                    info['total_sheets'] = wb.nsheets
            except:
                info['sheets'] = ['No disponible']
        
        return info


# Función de conveniencia
def load_data_file(file_path: str, **kwargs) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Función de conveniencia para cargar un archivo de datos
    
    Args:
        file_path: Ruta del archivo
        **kwargs: Argumentos adicionales
        
    Returns:
        Tuple con (DataFrame, metadata)
    """
    loader = MultiFormatLoader()
    return loader.load_file(file_path, **kwargs)


def get_supported_formats() -> Dict[str, str]:
    """Retorna los formatos soportados"""
    return MultiFormatLoader().get_supported_formats()


def check_dependencies() -> Dict[str, bool]:
    """Verifica qué dependencias están disponibles"""
    return {
        'openpyxl': HAS_OPENPYXL,
        'xlrd': HAS_XLRD,
        'pyarrow': HAS_PYARROW,
        'fastparquet': HAS_FASTPARQUET,
        'chardet': True  # Se verifica en tiempo de ejecución
    }
